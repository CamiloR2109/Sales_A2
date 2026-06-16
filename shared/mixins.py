from django.contrib import messages
from django.shortcuts import redirect


class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    
    Uso:
        class BrandDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
            ...
    
    ¿POR QUÉ?
    Porque solo el personal autorizado (staff) debe poder
    eliminar registros. Un usuario normal puede ver y crear,
    pero no borrar información importante del sistema.
    
    ¿CÓMO FUNCIONA?
    1. El usuario intenta acceder a una vista protegida
    2. dispatch() se ejecuta ANTES que la vista
    3. Si user.is_staff es False → redirige con mensaje de error
    4. Si user.is_staff es True → ejecuta la vista normalmente
    """

    # URL a donde redirigir si no es staff
    # Se puede sobreescribir en cada vista
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        """
        dispatch() es el primer método que se ejecuta en una CBV.
        Interceptamos aquí para verificar permisos ANTES de
        procesar la petición (GET o POST).
        """
        # Verificar si el usuario es staff
        if not request.user.is_staff:
            # Mostrar mensaje de error al usuario
            messages.error(request, self.staff_error_message)
            # Redirigir a la URL configurada
            return redirect(self.staff_redirect_url)

        # Si es staff, continuar con el flujo normal de la vista
        return super().dispatch(request, *args, **kwargs)


class ExportMixin:
    """
    Mixin para exportar querysets de ListView a PDF o Excel.
    Requiere que la vista defina:
    - export_fields: Lista de nombres de campos o rutas de atributos (ej: ['name', 'brand.name'])
    - export_headers: Lista de cabeceras correspondientes para las columnas.
    - export_filename: Nombre base del archivo descargable.
    """
    export_fields = []
    export_headers = []
    export_filename = "export"

    def get_export_fields(self):
        return self.export_fields

    def get_export_headers(self):
        return self.export_headers or self.get_export_fields()

    def get_export_filename(self, file_format):
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        return f"{self.export_filename}_{timestamp}.{file_format}"

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in ('pdf', 'excel'):
            self.object_list = self.get_queryset()
            if export_format == 'excel':
                return self.export_to_excel()
            elif export_format == 'pdf':
                return self.export_to_pdf()
        return super().get(request, *args, **kwargs)

    def resolve_field_value(self, obj, field_path):
        """
        Resuelve rutas de atributos con puntos (ej: 'brand.name').
        Maneja relaciones M2M (managers) uniendo sus elementos.
        Maneja booleanos traduciéndolos a Sí/No.
        """
        from django.db.models import Manager
        current_val = obj
        for attr in field_path.split('.'):
            if current_val is None:
                return ""
            val = getattr(current_val, attr, None)
            if callable(val) and not isinstance(val, Manager):
                try:
                    current_val = val()
                except TypeError:
                    current_val = val
            else:
                current_val = val

        if current_val is None:
            return ""

        if isinstance(current_val, Manager):
            return ", ".join(str(item) for item in current_val.all())

        if isinstance(current_val, bool):
            return "Sí" if current_val else "No"

        return current_val

    def export_to_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename[:30]

        headers = self.get_export_headers()
        ws.append(headers)

        # Estilo para cabecera (negrita, fondo oscuro, texto blanco)
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Agregar datos
        for obj in self.object_list:
            row = []
            for field in self.get_export_fields():
                row.append(self.resolve_field_value(obj, field))
            ws.append(row)

        # Ajuste automático del ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename("xlsx")}"'
        wb.save(response)
        return response

    def export_to_pdf(self):
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from django.http import HttpResponse
        from django.utils import timezone

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename("pdf")}"'

        # Usar orientación horizontal (landscape) para que quepan bien las columnas de las tablas
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter),
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
        )

        story = []
        styles = getSampleStyleSheet()

        # Estilos de título y subtítulo
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#212529'),
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#6C757D'),
            spaceAfter=20
        )

        title_text = f"Reporte de {self.export_filename.replace('_', ' ').title()}"
        story.append(Paragraph(title_text, title_style))
        
        timestamp_str = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
        story.append(Paragraph(f"Generado el: {timestamp_str} | Total de registros: {len(self.object_list)}", subtitle_style))

        # Preparar datos de la tabla
        headers = self.get_export_headers()
        table_data = [[Paragraph(h, ParagraphStyle('HStyle', parent=styles['Normal'], fontName='Helvetica-Bold', textColor=colors.white)) for h in headers]]

        cell_style = ParagraphStyle('CStyle', parent=styles['Normal'], fontSize=9, leading=11)
        for obj in self.object_list:
            row = []
            for field in self.get_export_fields():
                val = self.resolve_field_value(obj, field)
                row.append(Paragraph(str(val), cell_style))
            table_data.append(row)

        # Ancho utilizable en landscape letter (792 pt de ancho total - 60 pt de márgenes = 732 pt)
        num_cols = len(headers)
        col_width = 732.0 / num_cols if num_cols > 0 else 100

        t = Table(table_data, colWidths=[col_width] * num_cols)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#343A40')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DEE2E6')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))

        story.append(t)
        doc.build(story)
        return response

